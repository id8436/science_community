from django.contrib.auth.decorators import login_required #로그인이 있어야 가능함
from django.shortcuts import render, get_object_or_404, redirect, resolve_url, HttpResponse
import json
from django.utils import timezone  # 시간입력을 위해.
from _datetime import datetime,timedelta  # 시간데이터 활용을 위해
from django.contrib import messages  # 넌필드 오류를 반환하기 위한 것

from .forms import PostingForm, AnswerForm, CommentForm, ScoreForm
from .models import * #모델을 불러온다.
from django.db.models import Q, Count  # 검색을 위함. filter에서 OR조건으로 조회하기 위한 함수.(장고제공)
import openpyxl
from custom_account.models import Notification
from custom_account.views import notification_add

def tag_adding(request, posting):
    tags = request.POST.get('tag').split(',')
    for tag in tags:
        if not tag:
            continue
        else:
            tag = tag.strip()  # 문자열 양쪽에 빈칸이 있을 때 이를 제거한다.
            tag_, created = Tag.objects.get_or_create(name=tag)  # created엔 새로 만들어졌는지 여부가 True로 나오고, tag_엔 그 태그가 담긴다.
            posting.tag.add(tag_)  # 요건 특이하게 save()가 필요 없다.

def tag_info(request, tag_name):
    tag = get_object_or_404(Tag, name=tag_name)
    postings = tag.posting_set.all()
    context = {'tag':tag,
               'boards':postings}
    return render(request, 'boards/tag_info.html', context)
def tag_delete(request, posting_id, tag_name):
    '''게시글에서 태그를 지운다.'''
    posting = get_object_or_404(Posting, pk=posting_id)

    tag = get_object_or_404(Tag, name=tag_name)

@login_required()
def posting_like(request, posting_id):
    posting = get_object_or_404(Posting, pk=posting_id)
    context = {}
    if posting.like_users.filter(id=request.user.id).exists():
        posting.like_users.remove(request.user)  # 이미 추가되어 있다면 삭제한다.
        posting.like_count -= 1
        posting.save()  # save가 있어야 반영된다.
        context['like_check'] = False
    else:
        posting.like_users.add(request.user)  # 포스팅의 likeUser에 user를 더한다.
        posting.like_count += 1
        posting.save()
        context['like_check'] = True
    return HttpResponse(json.dumps(context), content_type='application/json')
@login_required()
def posting_dislike(request, posting_id):
    posting = get_object_or_404(Posting, pk=posting_id)
    context = {}
    if posting.dislike_users.filter(id=request.user.id).exists():
        posting.dislike_users.remove(request.user)  # 이미 추가되어 있다면 삭제한다.
        posting.dislike_count -= 1
        posting.save()  # save가 있어야 반영된다.
        context['like_check'] = False
    else:
        posting.dislike_users.add(request.user)  # 포스팅의 likeUser에 user를 더한다.
        posting.dislike_count += 1
        posting.save()
        context['like_check'] = True
    return HttpResponse(json.dumps(context), content_type='application/json')
@login_required()
def posting_interest(request, posting_id):
    posting = get_object_or_404(Posting, pk=posting_id)
    context = {}
    if posting.interest_users.filter(id=request.user.id).exists():
        posting.interest_users.remove(request.user)  # 이미 추가되어 있다면 삭제한다.
        posting.interest_count -= 1
        posting.save()  # save가 있어야 반영된다.
        context['posting_interest_check'] = False
    else:
        posting.interest_users.add(request.user)  # 포스팅의 likeUser에 user를 더한다.
        posting.interest_count += 1
        posting.save()
        context['posting_interest_check'] = True
    return HttpResponse(json.dumps(context), content_type='application/json')
@login_required()
def board_interest(request, board_id):
    board = get_object_or_404(Board, pk=board_id)
    context = {}
    if board.interest_users.filter(id=request.user.id).exists():
        board.interest_users.remove(request.user)  # 이미 추가되어 있다면 삭제한다.
        board.interest_count -= 1
        board.save()  # save가 있어야 반영된다.
        context['board_interest_check'] = False
    else:
        board.interest_users.add(request.user)  # 포스팅의 likeUser에 user를 더한다.
        board.interest_count += 1
        board.save()
        context['board_interest_check'] = True
    return HttpResponse(json.dumps(context), content_type='application/json')
@login_required()
def answer_create(request, posting_id):
    posting = get_object_or_404(Posting, pk=posting_id)
    if request.method == "POST":
        form = AnswerForm(request.POST)  # post를 통해 받은 폼.
        if form.is_valid():
            answer = form.save(commit=False)
            answer.posting = posting  # 상위 객체 연결.
            answer.author = request.user  # 추가한 속성 author 적용
            answer.create_date = timezone.now()
            answer.save()
            redirect_url = '{}#answer_{}'.format(request.META.get('HTTP_REFERER','/'), answer.id)
            notification_add(request, type=22, to_users=posting.interest_users.all(), message=posting.subject, url=redirect_url)
            return redirect(redirect_url)

    form = AnswerForm()
    context = {'posting': posting, 'form': form}
    return render(request, 'boards/detail.html', context)
    #폼이 적절하지 않다면 에러메시지와 함께 detail로 보낸다.

@login_required()
def answer_update(request, answer_id):
    answer = get_object_or_404(Answer, pk=answer_id)
    if request.user != answer.author:
        messages.error(request, '수정권한이 없습니다')
        return redirect('boards:detail', posting_id=answer.posting.id)

    if request.method == "POST":
        form = AnswerForm(request.POST, instance=answer)
        if form.is_valid():
            answer = form.save(commit=False)
            answer.author = request.user
            answer.modify_date = timezone.now()
            answer.save()
            return redirect('boards:detail', posting_id=answer.posting.id)
        else:
            form = AnswerForm(instance=answer)
        context = {'answer': answer, 'form': form}
    return render(request, 'answer_form.html', context)

@login_required()
def answer_delete(request, answer_id):
    answer = get_object_or_404(Answer, pk=answer_id)
    if request.user != answer.author:
        messages.error(request, '댓글삭제권한이 없습니다. 꼼수쓰지 마라;')
    else:
        answer.delete()
    return redirect('boards:detail', posting_id=answer.posting.id)

##############################################대댓글 관련 뷰##
@login_required()
def comment_create(request, answer_id):
    answer = get_object_or_404(Answer, pk=answer_id)
    if request.method == "POST":
        form = CommentForm(request.POST)
        if form.is_valid():
            comment = form.save(commit=False)
            comment.author = request.user
            comment.create_date = timezone.now()
            comment.answer = answer
            comment.save()
            redirect_url = '{}#comment_{}'.format(request.META.get('HTTP_REFERER','/'), comment.id)
            to_users = [answer.author]
            notification_add(request, type=32, to_users=to_users, message=answer.content, url=redirect_url)
            #return HttpResponse(json.dumps(context), content_type='application/json')
            return redirect('{}#answer_{}'.format(
                request.META.get('HTTP_REFERER','/'), answer.id))
            #return render(request, 'boards/comment_create_ajax.html', {'comment':comment})
    else:
        messages.error(request, '제대로 기입하쇼~')
        return redirect('boards:detail', posting_id=answer.posting.id)

@login_required()
def comment_update(request, comment_id):
    comment = get_object_or_404(Comment, pk=comment_id)
    if request.user != comment.author:
        messages.error(request, '댓글수정권한이 없습니다')
        return redirect('boards:detail', posting_id=comment.answer.posting.id)

    if request.method == "POST":
        form = CommentForm(request.POST, instance=comment)
        if form.is_valid():
            comment = form.save(commit=False)
            comment.author = request.user
            comment.modify_date = timezone.now()
            comment.save()
            return redirect('boards:detail', posting_id=comment.answer.posting.id)
    else:
        form = CommentForm(instance=comment)
    context = {'form': form}
    return render(request, 'comment_form.html', context)

@login_required()
def comment_delete(request, comment_id):
    comment = get_object_or_404(Comment, pk=comment_id)
    if request.user != comment.author:
        messages.error(request, '댓글삭제권한이 없습니다. 꼼수쓰지 마라;')
        return redirect('boards:detail', posting_id=comment.answer.posting.id)
    else:
        comment.delete()
    #return HttpResponse(json.dumps({}))
    #return redirect('boards:detail', posting_id=comment.answer.posting.id)

def subject_create(request, board_id):
    '''시험과목을 올린다.'''
    board = get_object_or_404(Board, pk=board_id)
    context = {}
    if request.method == 'POST':  # 포스트로 요청이 들어온다면... 글을 올리는 기능.
        box = request.POST.getlist('subject')
        sj_code = request.POST.getlist('sj_code')
        for i, tag in enumerate(box):
            if not tag:
                continue
            else:
                tag = tag.strip()  # 문자열 양쪽에 빈칸이 있을 때 이를 제거한다.
                tag_, created = Subject.objects.get_or_create(base_exam=board, name=tag, sj_code=sj_code[i])  # 과목 생성.
        return redirect('boards:board_detail', board_id=board_id)  # 작성이 끝나면 작성한 글로 보낸다.
    else:  # 포스트 요청이 아니라면.. form으로 넘겨 내용을 작성하게 한다.
        form = PostingForm()
        subjects = Subject.objects.filter(base_exam=board)  # 해당 보드에 해당하는 것을 가져온다.
        subjects = subjects.order_by('sj_code')  # 과목코드로 정렬.
        context['subjects'] = subjects
    context['form'] = form  # 폼에서 오류가 있으면 오류의 내용을 담아 create.html로 넘긴다.
    return render(request, 'boards/score/subject_create.html', context)

def subject_register(request, board_id):
    '''시험점수를 등록한다.'''
    board = get_object_or_404(Board, pk=board_id)
    context = {'board' : board}
    # 시험 주관기관에 해당하는 학생프로필이 있는지 찾기 위해.
    target_student = check.Check_student(request, board.school).in_school_and_none()

    if request.method == 'POST':  # 포스트로 요청이 들어온다면... 글을 올리는 기능.
        form = ScoreForm(request.POST)
        if form.is_valid():
            box = request.POST.getlist('subject_score')
            test_code = request.POST.get('test_code')
            profile, created = Exam_profile.objects.get_or_create(master=request.user, base_exam=board)  # 하나의 프로필만 단들게끔.
            profile.test_code = test_code
            if created:
                from boards.templatetags.board_filter import create_random_name
                profile.name = create_random_name(10)
            if target_student:  # 이미 등록되어 있는 학생계정이 있다면 다른 코드는 입력하지 못하게.
                profile.test_code = target_student.student_code
                # messages.error(request, '이미 등록된 정보가 있어 다른 코드는 입력할 수 없습니다.')
            profile.modify_num += 1  # 수정 할때마다 추가.
            profile.save()
            subjects = Subject.objects.filter(base_exam=board)
            for i, subject in enumerate(subjects):
                score = box[i]
                tag_, created = Score.objects.get_or_create(user=profile, base_subject=subject)  # 점수는 과목당 하나만 개설하게끔.
                tag_.score = score
                tag_.save()
                if target_student != None:
                    pass

            return redirect('boards:board_detail', board_id=board_id)  # 작성이 끝나면 작성한 글로 보낸다.
    else:  # 포스트 요청이 아니라면.. form으로 넘겨 내용을 작성하게 한다.
        form = ScoreForm()
    context['student'] = target_student
    return render(request, 'boards/score/subject_register.html', context)

def subject_download_excel_form(request, board_id):
    board = get_object_or_404(Board, pk=board_id)
    wb = openpyxl.Workbook()
    ws = wb.create_sheet('명단 form', 0)
    subjects = board.subject_set.all()
    ws['A1'] = '학번'
    ws['B1'] = '이름'
    for i, subject in enumerate(subjects):  # 과목명을 가장 첫행에 기입해넣는다.
        ws.cell(row=1, column=i+3).value = subject.name

    response = HttpResponse(content_type="application/vnd.ms-excel")
    response["Content-Disposition"] = 'attachment; filename=' + '명단양식' + '.xls'
    wb.save(response)

    return response

from school_report.view import check
def subject_upload_excel_form(request, board_id):
    if request.method == "POST":
        board = get_object_or_404(Board, pk=board_id)
        school = board.school
        if check.Check_teacher(request, board.school).in_school_and_none():
            pass
        else:
            messages.error(request, '이 기능은 관리자만이 가능합니다.')
            return redirect('boards:board_detail', board_id=board_id)

        uploadedFile = request.FILES["uploadedFile"]  # post요청 안의 name속성으로 찾는다.
        wb = openpyxl.load_workbook(uploadedFile, data_only=True)  # 파일을 핸들러로 읽는다.
        work_sheet = wb["명단 form"]  # 첫번째 워크시트를 사용한다.

        # 엑셀 데이터를 리스트 처리한다.
        work_sheet_data = []  # 전체 데이터를 담기 위한 리스트.
        for row in work_sheet.rows:  # 열을 순회한다.
            row_data = []  # 열 데이터를 담기 위한 리스트
            for cell in row:
                row_data.append(cell.value)  # 셀 값을 하나씩 리스트에 담는다.
            work_sheet_data.append(row_data)  # 워크시트 리스트 안에 열 리스트를 담아...
            # work_sheet_data[열번호][행번호] 형태로 엑셀의 데이터에 접근할 수 있게 된다.

        subject_info = work_sheet_data[0]  # 첫행은 과목을 생성하는 데 사용한다.
        subject_list = []  # 과목을 담을 리스트. 객체가 담긴다.
        for i in range(len(subject_info)-2):
            subject_name = subject_info[i+2]
            subject, created = Subject.objects.get_or_create(base_exam=board, name=subject_name)
            subject_list.append(subject)

        work_sheet_data = work_sheet_data[1:]  # 첫번째 행은 버린다.
        for data in work_sheet_data:  # 행별로 데이터를 가져온다.
            student_code = str(data[0])
            name = str(data[1])
            try:
                student = Student.objects.get(school=school, student_code=student_code, name=name)
            except Exception as e:
                print(e)
                messages.error(request, '수험자 정보에 이상이 있습니다. 기관에 먼저 등록하세요.')
                return redirect('boards:board_detail', board_id=board_id)

            if student.admin != None:  # 이 서비스를 이용하지 않는사람에겐 학생계정이 없어 문제가 생긴다.
                user = student.admin  # 계정 소유자.
                exam_profile, created = Exam_profile.objects.get_or_create(master=user, base_exam=board)  # 시험용 프로필.
            else:
                exam_profile, created = Exam_profile.objects.get_or_create(student=student, base_exam=board)
            if created:
                from boards.templatetags.board_filter import create_random_name
                exam_profile.name = create_random_name(10)
                exam_profile.save()
            for i, subject in enumerate(subject_list):
                score, created = Score.objects.get_or_create(user=exam_profile, base_subject=subject)
                score.real_score = data[i+2]  # 데이터로 들어온 점수를 넣어준다.
                score.save()
                board.official_check = True  # 공식 점수가 올라갔음을 의미.
                board.official_teacher = request.user
                board.save()

    return redirect('boards:board_detail', board_id=board_id)

