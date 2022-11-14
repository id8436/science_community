from django.contrib.auth.decorators import login_required #로그인이 있어야 가능함
from django.shortcuts import render, get_object_or_404, redirect, resolve_url, HttpResponse
import json
from django.utils import timezone  # 시간입력을 위해.
from _datetime import datetime,timedelta  # 시간데이터 활용을 위해
from django.contrib import messages  # 넌필드 오류를 반환하기 위한 것

from ..models import * #모델을 불러온다.
from django.db.models import Q, Count  # 검색을 위함. filter에서 OR조건으로 조회하기 위한 함수.(장고제공)
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Color
from school_report.view import check
from custom_account.models import Notification
from custom_account.views import notification_add
import numpy
import pandas as pd

def profile_create(request, board_id):
    base_exam = get_object_or_404(Board, pk=board_id)
    exam_profile, created = Exam_profile.objects.get_or_create(master=request.user, base_exam=base_exam)
    if created:
        from boards.templatetags.board_filter import create_random_name
        exam_profile.name = create_random_name(10)
    return redirect('boards:board_detail', board_id=base_exam.id)

def calculate_score(score_list):
    average = numpy.mean(score_list).round(2)
    variation = numpy.var(score_list).round(2)
    std = numpy.std(score_list).round(2)
    max_score = max(score_list)  # 최고점
    min_score = min(score_list)
    return [average, variation, std, max_score, min_score]

def statistical_of_score(request, subject):
    '''과목객체를 받아 통계데이터를 뱉어내는 함수.'''
    scores = subject.score_set.all()  # 과목 내의 점수들을 가져오고,
    code_list = []  # 학생코드 담을 것.
    score_list = [] # 점수 담을 것.
    try:  # 점수가 등록되지 않는 과목들이 있는 경우.
        if scores.last().real_score:  # 공식으로 등록된 점수가 있다면 공식 점수를...
            for score in scores:
                code = score.user.test_code
                code_list.append(code)
                score_list.append(score.real_score)
        else:
            for score in scores:
                code = score.user.test_code
                code_list.append(code)
                score_list.append(score.score)
    except:
        messages.error(request, str(subject) + '에 대해 등록된 점수가 없습니다.')

    # 동점자 수 구하기.
    same_count = score_list.copy()
    for i in range(len(score_list)):
        origin_score = same_count[i]
        same_count[i] = 0
        for score in score_list:
            if score == origin_score:
                same_count[i] += 1

    df = pd.DataFrame(code_list)
    df['score'] = score_list
    df['rank'] = df['score'].rank(method='min', ascending=False).astype(int)  # 소수점 없이 정수로 반환.
    df['rank'] = round(df['rank'] / df['rank'].count() * 100, 2)  # 랭크를 백분율로 바꾼다.
    df['same_rank'] = same_count
    df.set_index(0, inplace=True)
    return df
def result_main(request, board_id):
    board = get_object_or_404(Board, pk=board_id)
    context = {'board':board}
    subject_list = board.subject_set.filter(base_exam=board)  # 교과들을 불러와서...
    subject_data = {}  # 탬플릿에 보낼 과목정보.
    df_dict = {}  # 과목별 df를 모을 사전.
    subject_chart = {}  # 차트를 그릴 데이터를 담을 사전.
    for subject in subject_list:
        df = statistical_of_score(request, subject)
        # 점수 리스트를 구했으니, 이를 조작해 다양한 걸 얻을 수 있다.
        subject_info = calculate_score(df['score'])  # 해당 과목의 평균, 분산, 표준편차 등 데이터를 얻는다.
        subject_data[subject] = subject_info  # 교과와 교과정보를 한데 담아 보낸다.
        df_dict[subject] = df  # 아래에서 쓰기 위해 일단 저장.

        # 통계데이터 만들기
        max = subject_info[3]
        min = subject_info[4]
        interval_size = 10
        n = (max-min)/ interval_size
        data_dict = {}
        for i in range(interval_size):
            ceriterion_min = round(min + n*i, 2)
            ceriterion_max = round(min + n*(i+1), 2)
            interval_count = df.loc[(df['score'] > ceriterion_min) & (df['score'] <= ceriterion_max)].shape[0]  # 해당구간 데이터 세기.
            key_text = '{}초과, {}이하'.format(ceriterion_min, ceriterion_max)
            data_dict[key_text] = interval_count
        subject_chart[subject] = data_dict
    context['subject_data'] = subject_data
    context['subject_chart'] = subject_chart

    # 프로필 가져오기.
    exam_profile = Exam_profile.objects.get(master=request.user, base_exam=board)
    context['exam_profile'] = exam_profile
    studnet_code = exam_profile.test_code
    self_data = {}
    for subject in subject_list:
        subject_score_data = []
        df = df_dict[subject]  # 불러오기.
        info = df.loc[studnet_code]

        ## 본인의 점수를 담았으니, 각종 작업 수행.
        score = info['score']
        # 자신의 점수를 담는다.
        subject_score_data.append(score)
        # 표준점수 계산
        mean = subject_data[subject][0]
        std = subject_data[subject][2]
        std_score = (score - mean) / std
        subject_score_data.append(std_score)
        # 랭크데이터 계산
        rank = info['rank']
        same_count = int(info['same_rank'])
        rank_test = "상위 {}%({}명)".format(rank, same_count)
        subject_score_data.append(rank_test)

        self_data[subject] = subject_score_data
    context['self_data'] = self_data

    # 통계데이터 얻기.
    # for subject in subject_list:
    #     subject_score_data = []
    #     df = df_dict[subject]
    #     n =

    return render(request, 'boards/score/result/main.html', context)
def result_for_teacher(request, subject_id):
    context = {}
    subject = get_object_or_404(Subject, pk=subject_id)
    df = statistical_of_score(request, subject)
    context['dataframe'] = df

    return render(request, 'boards/score/result/main_for_teacher.html', context)

def subject_answer_info_form_download(request, subject_id):
    subject = get_object_or_404(Subject, pk=subject_id)
    wb = openpyxl.Workbook()
    ws = wb.create_sheet('명단 form', 0)
    ws['A1'] = subject.name  # 과목명을 가장 첫행에 기입해넣는다.
    ws['A3'] = '학번↓'
    ws['B1'] = '문항번호->'
    ws['B2'] = '배점->'
    ws['B3'] = '문항정답->'
    ws['B4'] = '이 열은 비우기'
    ws.column_dimensions['B'].width = 10

    for i in range(10):  # 기존 틀 생성.
        ws.cell(row=1, column=i + 3).value = i+1  # 문항번호.
        ws.cell(row=2, column=i + 3).value = 0  # 문항배점칸.
        ws.cell(row=3, column=i + 3).value = 0   # 정답칸.

    # 기존 입력 데이터 반영.
    answers = json.loads(subject.right_answer)
    for i, answer in enumerate(answers):
        ws.cell(row=3, column=i+3).value = answer
    distributions = json.loads(subject.distribution)
    for i, distribution in enumerate(distributions):
        ws.cell(row=2, column=i+3).value = distribution
    scores = subject.score_set.all()
    for i, score in enumerate(scores):
        ws.cell(row=4 + i, column=1).value = score.user.test_code  # 학번 반영.
        for j, answer in enumerate(answers):
            ws.cell(row=4+i, column=j+3).value = answer  # 학생의 답 써넣기.

    # 색 채우기
    black_fill = PatternFill(fill_type='solid', fgColor=Color('000000'))
    yellow_fill = PatternFill(fill_type='solid', fgColor=Color('ffff99'))
    red_fill = PatternFill(fill_type='solid', fgColor=Color('ff9999'))
    for cell in ws["B"]:
        cell.fill = black_fill
    for cell in ws["1"]:
        cell.fill = yellow_fill
    for cell in ws["2:2"]:
        cell.fill = red_fill
    for cell in ws["3"]:
        cell.fill = yellow_fill
    for cell in ws["A:A"]:
        cell.fill = yellow_fill


    response = HttpResponse(content_type="application/vnd.ms-excel")
    response["Content-Disposition"] = 'attachment; filename=' + '명단양식' + '.xls'
    wb.save(response)

    return response


def subject_answer_info_form_upload(request, subject_id):
    subject = get_object_or_404(Subject, pk=subject_id)
    board = subject.base_exam
    school = subject.base_exam.school
    if check.Check_teacher(request, school).in_school_and_none() and request.method == "POST":
        pass
    else:
        return check.Check_teacher(request, school).redirect_to_school()
    uploadedFile = request.FILES["uploadedFile"]  # post요청 안의 name속성으로 찾는다.
    wb = openpyxl.load_workbook(uploadedFile, data_only=True)  # 파일을 핸들러로 읽는다.
    work_sheet = wb["명단 form"]  # 첫번째 워크시트를 사용한다.

    # 엑셀 데이터를 리스트 처리한다.
    work_sheet_data = []  # 전체 데이터를 담기 위한 리스트.
    for row in work_sheet.rows:  # 열을 순회한다.
        row_data = []  # 열 데이터를 담기 위한 리스트
        for cell in row:
            row_data.append(cell.value)  # 셀 값을 하나씩 리스트에 담는다.
        work_sheet_data.append(row_data)  # 워크시트 리스트 안에 열 리스트를 담아...
        # work_sheet_data[열번호][행번호] 형태로 엑셀의 데이터에 접근할 수 있게 된다.
    # meta data.
    meta = work_sheet_data[0]
    if subject.name == meta[0]:
        pass
    else:
        messages.error(request,'과목정보가 다릅니다. 등록하고자 한 과목:' + subject.name + ', 등록한 과목:'+str(meta[0]))
        return redirect('boards:board_detail', board_id=board.id)

    answer_info = work_sheet_data[2]  # 3행은 정답정보.
    # 과목객체에 정답정보 담기.
    answer_info = work_sheet_data[2]  # 3행은 정답정보.
    right_answer_list = []
    for i in range(len(answer_info)-2):
        answer = answer_info[i+2]
        if answer == None:
            answer = 0
        right_answer_list.append(answer)  # 정답 순서대로 담는다.
    subject.right_answer = json.dumps(right_answer_list)
    # 배점정보 담기.
    distribution_info = work_sheet_data[1]  # 2행은 배점정보.
    distribution_list = []
    for i in range(len(distribution_info)-2):
        distribution = distribution_info[i+2]
        if distribution == None:
            distribution = 0
        distribution_list.append(distribution)  # 정답 순서대로 담는다.
    subject.distribution = json.dumps(distribution_list)
    subject.save()  # 정보를 담고 저장.

    # 학생 정보 저장.
    work_sheet_data = work_sheet_data[3:]  # 1~3번째 행은 버린다.(메타데이터)
    for data in work_sheet_data:  # 행별로 데이터를 가져온다.
        if data[0] == None:  # 위의 행을 비우고 올린 경우가 있다. 이런 경우엔 지나가주자.
            continue
        student_code = str(data[0])
        try:
            student = Student.objects.get(school=school, student_code=student_code)
        except Exception as e:
            messages.error(request, str(student_code) +'수험자 정보에 이상이 있습니다. 수험자를 기관에 먼저 등록하세요.')
            return redirect('boards:board_detail', board_id=board.id)

        if student.admin != None:  # 이 서비스를 이용하지 않는사람에겐 학생계정이 없어 문제가 생긴다.
            user = student.admin  # 계정 소유자.
            exam_profile, created = Exam_profile.objects.get_or_create(master=user, base_exam=board)  # 시험용 프로필.
            exam_profile.student = student
        else:
            exam_profile, created = Exam_profile.objects.get_or_create(student=student, base_exam=board)
        if created:
            from boards.templatetags.board_filter import create_random_name
            exam_profile.name = create_random_name(10)  # 새로 생성되었다면 이름 배정조치.
        exam_profile.test_code = student_code
        exam_profile.save()

        score, created = Score.objects.get_or_create(user=exam_profile, base_subject=subject)  # 점수 생성.
        user_answer = []  # 사용자의 답을 담기 위한 리스트.
        for i in range(len(data)-2):
            answer = data[i+2]
            user_answer.append(answer)
        score.answer = json.dumps(user_answer)
        # 점수 계산.
        try:  # 배점정보가 있을 때.
            test = distribution_list[0]
            result_score = 0
            for right, answer, distribution in zip(right_answer_list, user_answer, distribution_list):
                if right == answer:
                     result_score += float(distribution)
            score.real_score = result_score
        except:
            pass
        score.save()  # 해당 학생의 답변을 저장하고 닫는다.
    messages.info(request, '등록 성공')
    board.official_check = True
    board.official_teacher = request.user
    board.save()

    return redirect('boards:board_detail', board_id=board.id)

def show_answer(request, score_id):
    # 나중에 과목별로 답변을 보게 하면 좋겠네. 한번에 볼 수 있게.
    score = Score.objects.get(id=score_id)
    if request.user == score.user.master:  # 당사자일 때에만 읽게 한다.
        context = {}
        decoder = json.decoder.JSONDecoder()  # 디코더객체 설정.
        subject = score.base_subject
        distribution = decoder.decode(subject.distribution)
        right_answer = decoder.decode(subject.right_answer)
        user_anser = decoder.decode(score.answer)
        context['answer_info'] = zip(distribution, right_answer, user_anser)  # 이중for문을 위하여~
        #context['user_answer'] = decoder.decode(score.answer)
        context['subject'] = score.base_subject
        return render(request, 'boards/score/result/show_answer.html', context)
    else:
        messages.error(request, '자신의 정답만 확인할 수 있습니다.')
    return redirect('boards:board_detail', board_id=score.base_subject.base_exam.id)

def show_answer_for_teacher(request, subject_id):
    subject = Subject.objects.get(id=subject_id)
    context = {}
    if check.Check_teacher(request, subject.base_exam.school).in_school_and_none():
        pass
    else:
        messages.error(request, '꼼수 쓰지 마라.')
        return redirect('boards:board_detail', board_id=subject.base_exam.id)
    scores = subject.score_set.all()
    decoder = json.decoder.JSONDecoder()  # 디코더객체 설정.
    # 정답 담기.
    if subject.right_answer:
        right_answer = decoder.decode(subject.right_answer)
        context['right_answer'] = right_answer
    # 배점 담기.
    if subject.distribution:
        distributions = decoder.decode(subject.distribution)
        context['distributions'] = distributions
    # 학생정보 담기.
    student_answer_info = {}
    for score in scores:
        try:  # 외부인이 계정 없이 임의로 등록한 거라면... 학생계정에 매칭이 되지 않는다. 그냥 무시.
            student = score.user.student.student_code
            info = [score.real_score]
            try:  # 정답에 대한 정보가 없으면 에러 난다.
                answer = decoder.decode(score.answer)  # 답을 리스트로.
                info.append(answer)  # 개별 점수와 정답정보를 담는다.
            except:
                pass
            student_answer_info[student] = info  # 사전 안에 리스트로 담는다.
        except:
            pass
    context['student_answer_info'] = student_answer_info
    return render(request, 'boards/score/result/show_answer_for_teacher.html', context)

