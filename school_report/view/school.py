from django.shortcuts import render, redirect
from .. import models  # 모델 호출.
from ..forms import HomeroomForm, SchoolForm, SubjectForm, SchoolLostItemForm
from django.contrib import messages
from custom_account.decorator import custom_login_required as login_required
import openpyxl
from . import check
from django.http import HttpResponse
import random
from boards.models import Board, Board_category, Board_name
from .for_api import SchoolMealsApi
from django.utils.encoding import escape_uri_path
from django.http import JsonResponse, HttpResponseForbidden
from django.contrib.auth.models import User
import datetime

def main(request, school_id):
    context = {}
    school = get_object_or_404(models.School, pk=school_id)
    context['school'] = school
    context['announcement_list'] = school.announcebox.announcement_set.order_by('-create_date')
    context['homework_list'] = school.homeworkbox.homework_set.order_by('-create_date')
    # 교사여부.
    teacher = check.Teacher(user=request.user, school=school, request=request).in_school_and_none()
    context['teacher'] = teacher
    # 학생여부.
    student = check.Student(user=request.user, school=school, request=request).in_school_and_none()
    context['student'] = student
    # 소속교실, 소속교과만 가져온다.
    if teacher:  # 교사인 경우엔 다.
        context['homeroom_list'] = school.homeroom_set.all().order_by('name')
        context['subject_list'] = school.subject_set.all().order_by('subject_name')
        context['classroom_list'] = school.classroom_set.all().order_by('homeroom__name', 'name')
    if student:
        homeroom_list = student.homeroom.all().order_by('name')
        context['homeroom_list'] = homeroom_list
        classroom_list = school.classroom_set.filter(homeroom__in=homeroom_list).order_by('base_subject').select_related(
            'homeroom')
        # subject_list = []
        # for homeroom in homeroom_list:
        #     subjects = school.subject_set.filter(homeroom=homeroom).order_by('subject_name')
        #     subject_list.append(subjects)
        context['classroom_list'] = classroom_list
    # 시험문제목록.
    category = Board_category.objects.get(id=6)
    context['category'] = category  # 게시판 생성 때 카테고리 아이디도 필요해서.
    exam_list = Board.objects.filter(school=school, category=category)
    context['exam_list'] = exam_list
    return render(request, 'school_report/school/main.html', context)
def list(request):
    context = {}
    school_list = models.School.objects.order_by('-id')  # 생성순서의 반대.
    #-----검색기능
    keyword = request.GET.get('keyword', '')  # 검색어를 받는다.
    if keyword != '':  # 검색어가 있다면
        result = []  # 검색결과를 담기 위한 리스트를 만든다.
        keywords = keyword.split(' ')  # 공백이 있는 경우 나눈다.
        for kw in keywords:  # 띄어쓰기로 검색을 하는 경우가 많으니까, 다 찾아줘야지.
            result += school_list.filter(name__icontains=kw) # 제목검색
    school_list = result
    context['board_list'] = school_list
    return render(request, 'school_report/school/list.html', context)

def name_trimming(name):
    instr = name
    outstr = ''
    for i in range(0, len(instr)):  # 문자열 내 공백 없애기.
        if instr[i] != ' ':
            outstr += instr[i]
    return outstr

@login_required()
def school_create(request):
    context = {}
    if request.method == 'POST':  # 포스트로 요청이 들어온다면... 글을 올리는 기능.
        form = SchoolForm(request.POST)  # 폼을 불러와 내용입력을 받는다.
        if form.is_valid():  # 문제가 없으면 다음으로 진행.
            school = form.save(commit=False)  # commit=False는 저장을 잠시 미루기 위함.(입력받는 값이 아닌, view에서 다른 값을 지정하기 위해)
            school.master = request.user  # 추가한 속성 author 적용
            name = name_trimming(school.name)  # 이름에서 공백 제거해 적용.
            school.name = name
            # 게시판 생성 및 연동.
            board_name = str(school.name) + " 교직원 게시판"
            board_name, _ = Board_name.objects.get_or_create(name=board_name)  # 이름객체를 생성한다.(이거 은근 불편하네;; 역대 게시판을 모을 수 있다는 점에선 좋지만... 검색해도 될듯?)
            category = Board_category.objects.get(pk=7)  # 교직원게시판의 카테고리.
            school.save()  # 객체 생성을 해야 게시판을 만들 수 있다.
            board, _ = Board.objects.get_or_create(board_name=board_name, category=category, enter_year=school.year, author=request.user, school=school)
            school.teacher_board_id = board.id  # 게시판 지정.
            school.save()
            # 과제박스 생성.
            homework_box, created = models.HomeworkBox.objects.get_or_create(school=school)
            announce_box, created = models.AnnounceBox.objects.get_or_create(school=school)
            return redirect('school_report:school_main', room_id=school.id)
    else:  # 포스트 요청이 아니라면.. form으로 넘겨 내용을 작성하게 한다.
        form = SchoolForm()
    context['form'] = form  # 폼에서 오류가 있으면 오류의 내용을 담아 create.html로 넘긴다.
    return render(request, 'school_report/school/school_create.html', context)
@login_required()
def school_modify(request, school_id):
    '''학교 관련 정보 수정.(게시판 사용여부 포함)'''
    school = get_object_or_404(models.School, pk=school_id)
    if request.user != school.master:
        messages.error(request, '수정권한이 없습니다')
        return redirect('school_report:school_main', school_id=school.id)
    if request.method == "POST":
        form = SchoolForm(request.POST, instance=school)  # 받은 내용을 객체에 담는다. instance에 제대로 된 걸 넣지 않으면 새로운 인스턴스를 만든다.
        if form.is_valid():
            school = form.save(commit=False)
            name = name_trimming(school.name)  # 이름에서 공백 제거해 적용.
            school.name = name
            # 기본적으로 게시판 사용 중단 후, 체크된 것 보고 진행.
            school.is_suggestion_board_active = False
            school.is_lost_item_board_active = False
            print(request.POST.getlist('boards_checks'))
            for board in request.POST.getlist('boards_checks'):
                if board == 'lost_item_board':
                    lost_item_board, _ = models.LostItemBoard.objects.get_or_create(school=school)
                    school.is_lost_item_board_active = True
                if board == 'suggestion_board':
                    sugg_board, _ = models.SuggestionBoard.objects.get_or_create(school=school)
                    school.is_suggestion_board_active = True
            school.save()
            return redirect('school_report:school_main', school_id=school.id)
    else:  # GET으로 요청된 경우.
        form = SchoolForm(instance=school)  # 해당 모델의 내용을 가져온다!
        # 태그를 문자열화 하여 form과 함께 담는다.
    context = {'form': form, 'school': school}
    return render(request, 'school_report/school/school_modify.html', context)

def create_performance_score(request, subject_id):
    subject = get_object_or_404(models.Subject, pk=subject_id)
    category = get_object_or_404(Board_category, pk=6)  # 점수게시판 생성을 위해.
    context = {}
    school = subject.school
    board_name_str = str(school.name) +'/'+ str(subject.subject_name) +'/'+ '수행평가'
    board_name, _ = Board_name.objects.get_or_create(name=board_name_str)
    board, _ = Board.objects.get_or_create(category=category, author=subject.master.admin, board_name=board_name, enter_year=school.year,
                                           school=school)
    return redirect('boards:board_detail', board.id)
@login_required()
def download_excel_form(request, school_id):
    '''교사 명단 폼.'''
    school = get_object_or_404(models.School, pk=school_id)
    wb = openpyxl.Workbook()
    ws = wb.create_sheet('명단 form', 0)
    ws['A1'] = '(선택)교사코드'
    ws['B1'] = '이름'
    ws['C1'] = '가입인증코드(없으면 랜덤 6개)'
    ws['D1'] = '학년'
    ws['E1'] = '반'
    ws['F1'] = '(선택)교실이름'
    #teachers = school.teacher_set.all()
    teachers = models.Profile.objects.filter(school=school, position='teacher')
    code_line = 'A'
    name_line = 'B'  # 이름 담을 라인.
    confirm_code_line = 'C'
    grade_line = 'D'
    cl_num_line = 'E'
    homeroom_name_line = 'F'
    for i, teacher in enumerate(teachers):
        num = str(i + 2)
        ws[code_line + num] = teacher.code
        ws[name_line + num] = teacher.name
        ws[confirm_code_line + num] = teacher.confirm_code
        try:
            ws[grade_line + num] = teacher.homeroom.last().grade  # 마지막 객체의 속성을 가져오는데 None이면 에러가 뜬다.
            ws[cl_num_line + num] = teacher.homeroom.last().cl_num
        except:
            pass
        try:
            ws[homeroom_name_line + num] = teacher.homeroom.last().name
        except:
            pass
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{escape_uri_path("교사명단.xlsx")}"'
    wb.save(response)

    return response


@login_required()
def upload_excel_form(request, school_id):
    context = {}
    if request.method == "POST":
        school = get_object_or_404(models.School, pk=school_id)
        context['school'] = school
        uploadedFile = request.FILES["uploadedFile"]  # post요청 안의 name속성으로 찾는다.
        wb = openpyxl.load_workbook(uploadedFile, data_only=True)  # 파일을 핸들러로 읽는다.
        work_sheet = wb.worksheets[0]  # 첫번째 워크시트를 사용한다.

        # 엑셀 데이터를 리스트 처리한다.
        work_sheet_data = []  # 전체 데이터를 담기 위한 리스트.
        for row in work_sheet.rows:  # 열을 순회한다.
            row_data = []  # 열 데이터를 담기 위한 리스트
            for cell in row:
                row_data.append(cell.value)  # 셀 값을 하나씩 리스트에 담는다.
            work_sheet_data.append(row_data)  # 워크시트 리스트 안에 열 리스트를 담아...
            # work_sheet_data[열번호][행번호] 형태로 엑셀의 데이터에 접근할 수 있게 된다.
        work_sheet_data = work_sheet_data[1:]  # 첫번째 행은 버린다.

        if request.user == school.master:
            for data in work_sheet_data:  # 행별로 데이터를 가져온다.
                profile_code = data[0]
                name = data[1]
                confirm_code = data[2]
                grade = data[3]
                cl_num = data[4]
                homeroom_name = data[5]
                profile, created = models.Profile.objects.get_or_create(name=name, code=profile_code, school=school, position='teacher')
                if not confirm_code:  # 이런 방식으로 받으면 문자열로 들어온다;
                    # 2개 데이터만 들어온 경우로, 코드정보가 없으면 랜덤으로 배정.
                    profile.confirm_code = random.randint(100000, 999999)  # 코드 지정.
                else:
                    profile.confirm_code = confirm_code
                profile.save()
                # 담임교실 배정 및 생성.
                if grade and cl_num:  # 학년, 반이 다 들어온 경우.
                    homeroom, cteated = models.Homeroom.objects.get_or_create(grade=grade, cl_num=cl_num, school=school)
                    homeroom_name_ = f'{grade}학년 {cl_num}반'
                    homeroom.name = homeroom_name_  # 학년,반으로 홈룸이름 지정.
                    homeroom.master_profile = profile
                    if homeroom_name:  # 홈룸네임이 있는 경우 덮어쓰기.
                        homeroom.name = homeroom_name
                elif homeroom_name:  # 홈룸네임만 있는 경우.
                    homeroom, cteated = models.Homeroom.objects.get_or_create(name=homeroom_name, school=school)
                    homeroom.master_profile = profile
                else:
                    continue  # 저장과정으로 넘어가지 않고 다음으로.
                homeroom.save()


    return redirect('school_report:teacher_assignment', school_id=school_id)  # 필요에 따라 렌더링.




@login_required()
def teacher_code_input(request, school_id):
    school = get_object_or_404(models.School, pk=school_id)
    context = {'school': school}
    if request.method == 'POST':  # 포스트로 요청이 들어온다면...
        confirm_code = request.POST.get('code')
        try:
            teacher = models.Profile.objects.filter(school=school, confirm_code=confirm_code)[0]  # 해당 계정 중 1번째.
            if teacher.obtained == True:
                messages.error(request, '이미 누군가 등록한 프로필입니다.')
                return render(request, 'school_report/school/teacher_code_input.html', context)
            # 이상 없으면 확인 페이지로 이동시킨다.
            return redirect('school_report:teacher_code_confirm', teacher_id=teacher.id)
        except:
            messages.error(request, '코드 해당계정이 없습니다.')
    return render(request, 'school_report/school/teacher_code_input.html', context)

def teacher_code_confirm(request, teacher_id):
    teacher = get_object_or_404(models.Profile, pk=teacher_id)
    context={'teacher':teacher}
    if request.method == 'POST':  # 포스트로 요청이 들어온다면...
        confirm_code = request.POST.get('code')
        if confirm_code == teacher.confirm_code:
            teacher.admin = request.user
            teacher.obtained = True
            teacher.confirm_code = None
            teacher.save()
            # request.user.teacher = teacher  # 계정에 등록.
            # request.user.save()  # 이것도 저장 해주어야 해.
            messages.info(request, '인증에 성공하였습니다.')
            return redirect('school_report:main')
        else:
            messages.error(request, '코드가 안맞는데요;')
            return render(request, 'school_report/school/teacher_code_confirm.html', context)

    return render(request, 'school_report/school/teacher_code_confirm.html', context)
def teacher_delete(request, teacher_id):
    teacher = get_object_or_404(models.Profile, pk=teacher_id)
    school = teacher.school
    if school.master == request.user:
        messages.error(request, "삭제하였습니다.")
        teacher.delete()
        return redirect(request.META.get('HTTP_REFERER', None))  # 이전 화면으로 되돌아가기.
    else:
        messages.error(request, '관리자만이 가능합니다.')
        return redirect(request.META.get('HTTP_REFERER', None))  # 이전 화면으로 되돌아가기.

@login_required()
def teacher_assignment(request, school_id):
    '''관리자가 보는 교사 명단'''
    school = get_object_or_404(models.School, pk=school_id)
    context = {'school': school}
    if school.master == request.user:
        teacher_list_resistered = models.Profile.objects.filter(school=school, obtained=True, position='teacher')  # 학교 내에 등록된 프로필만 가져온다.
        context['teacher_list_resistered'] = teacher_list_resistered

        teacher_list_unresistered = models.Profile.objects.filter(school=school, obtained=False, position='teacher')  # 등록 안한 사람만 반환.
        context['teacher_list_unresistered'] = teacher_list_unresistered
        return render(request, 'school_report/school/assignment.html', context)

# def school_profile_delete(request, profile_id):
#     '''다른 데 기능을 구현해서 쓸모 없어짐.'''
#     profile = get_object_or_404(models.Profile, pk=profile_id)
#     school = profile.school
#     if school.master == request.user:
#         messages.error(request, "삭제하였습니다.")
#         profile.delete()
#         return redirect(request.META.get('HTTP_REFERER', None))  # 이전 화면으로 되돌아가기.
#     else:
#         messages.error(request, '관리자만이 가능합니다.')
#         return redirect(request.META.get('HTTP_REFERER', None))  # 이전 화면으로 되돌아가기.

############################## 학생 관련.
def student_assignment(request, type, baseRoom_id):
    if type == 'school':
        school = get_object_or_404(models.School, pk=baseRoom_id)
        context = {'baseRoom': school}
        if school.master == request.user:
            student_list_resistered = models.Profile.objects.filter(school=school, obtained=True, position='student')  # 학교 내에 등록된 프로필만 가져온다.
            context['student_list_resistered'] = student_list_resistered

            student_list_unresistered = models.Profile.objects.filter(school=school, obtained=False, position='student')  # 등록 안한 사람만 반환.
            context['student_list_unresistered'] = student_list_unresistered
            return render(request, 'school_report/school/student_assignment.html', context)
    if type == 'homeroom':
        homeroom = get_object_or_404(models.Homeroom, pk=baseRoom_id)
        context = {'baseRoom': homeroom}
        if check.Teacher(user=request.user, homeroom=homeroom).in_homeroom_and_none:
            resistered = models.Profile.objects.filter(homeroom=homeroom, obtained=True)
            context['student_list_resistered'] = resistered
            unresistered = models.Profile.objects.filter(homeroom=homeroom, obtained=False)  # 등록 안한 사람만 반환.
            context['student_list_unresistered'] = unresistered
            return render(request, 'school_report/school/student_assignment.html', context)
        else:
            messages.error(request, '당신은 학급의 관리자가 아닙니다.')
            return redirect('school_report:homeroom_main', homeroom_id=baseRoom_id)
@login_required()
def student_download_excel_form(request, type, baseRoom_id):
    if type == 'school':
        school = get_object_or_404(models.School, pk=baseRoom_id)
        if school.master == request.user:
            pass
        else:
            messages.error(request, '부정접근.')
            return redirect(request.META.get('HTTP_REFERER', None))  # 이전 화면으로 되돌아가기.
        wb = openpyxl.Workbook()
        ws = wb.create_sheet('명단 form', 0)
        ws['A1'] = '학번'
        ws['B1'] = '이름'
        ws['C1'] = '가입인증코드(없으면 랜덤 6개)'
        ws['D1'] = '학년(선택사항)'
        ws['E1'] = '반(선택사항)'
        students = models.Profile.objects.filter(school=school, position='student')  # 학교에 속한 학생.
        a = 'A'  # 학번쓰는 라인.
        b = 'B'  # 이름쓰는 라인.
        c = 'C'  # 코드 쓰는 라인.
        d = 'D'  # 학년
        e = 'E'  # 반
        for i, student in enumerate(students):
            num = str(i + 2)
            ws[a + num] = student.code
            ws[b + num] = student.name
            if student.obtained:
                ws[c + num] = '등록함'
            else:
                ws[c + num] = student.confirm_code  # 가입인증코드.
            if student.homeroom:
                homeroom = student.homeroom.all().first()  # 여러 학급에 등록되어 있을 수 있으니.(그래도 기본적으로 1학급.)
                ws[d + num] = homeroom.grade
                ws[e + num] = homeroom.cl_num
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        file_name = f'학생명단({school}).xlsx'
        response['Content-Disposition'] = f'attachment; filename="{escape_uri_path(file_name)}"'
        wb.save(response)

        return response
    if type == 'homeroom':
        homeroom = get_object_or_404(models.Homeroom, pk=baseRoom_id)
        if check.Teacher(user=request.user, homeroom=homeroom).in_homeroom_and_none:
            pass
        else:
            return check.Teacher.redirect_to_homeroom()
        # students = homeroom.student_set.all()
        students = models.Profile.objects.filter(homeroom=homeroom)  # 학급에 속한 학생.
        wb = openpyxl.Workbook()
        ws = wb.create_sheet('명단 form', 0)
        ws['A1'] = '수험코드(학번)'
        ws['B1'] = '이름'
        ws['C1'] = '가입인증코드(없으면 랜덤 6개)'
        a = 'A'  # 번호쓰는 라인.
        b = 'B'  # 이름쓰는 라인.
        c = 'C'  # 코드 쓰는 라인.
        for i, student in enumerate(students):
            num = str(i + 2)
            ws[a + num] = student.code
            ws[b + num] = student.name
            if student.obtained:
                ws[c + num] = '등록함'
            else:
                ws[c + num] = student.confirm_code  # 가입인증코드.
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        file_name = f'학생명단({homeroom}).xlsx'
        response['Content-Disposition'] = f'attachment; filename="{escape_uri_path(file_name)}"'
        # response = HttpResponse(content_type="application/vnd.ms-excel")  # 25년 6월 넘어가면 없애자.
        # response["Content-Disposition"] = 'attachment; filename=' + '명단양식' + '.xls'
        wb.save(response)
        return response


@login_required()
def student_upload_excel_form(request, type, baseRoom_id):
    '''학생명단 업로드'''
    context = {}
    if request.method == "POST":
        if type == 'school':
            school = get_object_or_404(models.School, pk=baseRoom_id)
            if request.user == school.master:
                pass
            else:
                messages.error(request, '이 기능은 관리자만이 가능합니다.')
                return check.Teacher(school=school).redirect_to_school()
            school = get_object_or_404(models.School, pk=baseRoom_id)
            context['baseRoom'] = school

            uploadedFile = request.FILES["uploadedFile"]  # post요청 안의 name속성으로 찾는다.
            wb = openpyxl.load_workbook(uploadedFile, data_only=True)  # 파일을 핸들러로 읽는다.
            work_sheet = wb.worksheets[0]  # 첫번째 워크시트를 사용한다.

            # 엑셀 데이터를 리스트 처리한다.
            work_sheet_data = []  # 전체 데이터를 담기 위한 리스트.
            for row in work_sheet.rows:  # 열을 순회한다.
                row_data = []  # 열 데이터를 담기 위한 리스트
                for cell in row:
                    row_data.append(cell.value)  # 셀 값을 하나씩 리스트에 담는다.
                work_sheet_data.append(row_data)  # 워크시트 리스트 안에 열 리스트를 담아...
                # work_sheet_data[열번호][행번호] 형태로 엑셀의 데이터에 접근할 수 있게 된다.
            work_sheet_data = work_sheet_data[1:]  # 첫번째 행은 버린다.

            for data in work_sheet_data:  # 행별로 데이터를 가져온다.
                student_code = data[0]
                name = data[1]
                confirm_code = data[2]
                grade = data[3]
                cl_num = data[4]

                profile, created = models.Profile.objects.get_or_create(school=school, name=name, code=student_code, position='student')
                # if created:
                #     exam_profiles = student.exam_profile_set.all()
                #     for profile in exam_profiles:
                #         profile.master = student.admin
                #         profile.save()  # 시험등록 때 계정이 없던 사람은 시험프로필을 학생에 연결해두었으므로 이를 계정에 직접 연결해준다.

                # 학급정보가 있다면 만들어버리기.
                if not confirm_code:  # 이런 방식으로 받으면 문자열로 들어온다;
                    # 2개 데이터만 들어온 경우로, 코드정보가 없으면 랜덤으로 배정.
                    profile.confirm_code = random.randint(100000, 999999)  # 코드 지정.
                else:
                    profile.confirm_code = confirm_code
                if grade and cl_num:  # 학년, 반이 다 들어온 경우.
                    homeroom, cteated = models.Homeroom.objects.get_or_create(grade=grade, cl_num=cl_num, school=school)
                    profile.homeroom.add(homeroom)
                profile.save()
            messages.info(request, '반영 완료.')
            return redirect('school_report:student_assignment', type='school', baseRoom_id=baseRoom_id)

        if type == 'homeroom':
            homeroom = get_object_or_404(models.Homeroom, pk=baseRoom_id)
            if check.Teacher(user=request.user, homeroom=homeroom).in_homeroom_and_none:
                pass
            else:
                return check.Teacher.redirect_to_homeroom()
            context['baseRoom'] = homeroom
            uploadedFile = request.FILES["uploadedFile"]  # post요청 안의 name속성으로 찾는다.
            wb = openpyxl.load_workbook(uploadedFile, data_only=True)  # 파일을 핸들러로 읽는다.
            work_sheet = wb["명단 form"]  # 첫번째 워크시트를 사용한다.

            # 엑셀 데이터를 리스트 처리한다.
            work_sheet_data = []  # 전체 데이터를 담기 위한 리스트.
            for row in work_sheet.rows:  # 열을 순회한다.
                row_data = []  # 열 데이터를 담기 위한 리스트
                for cell in row:
                    row_data.append(cell.value)  # 셀 값을 하나씩 리스트에 담는다.
                work_sheet_data.append(row_data)  # 워크시트 리스트 안에 열 리스트를 담아...
                # work_sheet_data[열번호][행번호] 형태로 엑셀의 데이터에 접근할 수 있게 된다.
            work_sheet_data = work_sheet_data[1:]  # 첫번째 행은 버린다.

            for data in work_sheet_data:  # 행별로 데이터를 가져온다.
                student_code = data[0]
                name = data[1]
                confirm_code = data[2]
                student, created = models.Profile.objects.get_or_create(school=homeroom.school, name=name,
                                                                        # 전학생도 고려해야 하니, 생성도 허용한다.
                                                                        code=student_code, position='student')
                student.homeroom.add(homeroom)
                if not confirm_code:  # 이런 방식으로 받으면 문자열로 들어온다;
                    # 2개 데이터만 들어온 경우로, 코드정보가 없으면 랜덤으로 배정.
                    student.confirm_code = random.randint(100000, 999999)  # 코드 지정.
                else:
                    student.confirm_code = confirm_code
                if created:
                    student.confirm_code = random.randint(100000, 999999)  # 코드 지정.
                student.save()
            messages.info(request, '반영 완료.')
            return redirect('school_report:student_assignment', type='homeroom', baseRoom_id=baseRoom_id)

    return redirect(request.META.get('HTTP_REFERER', None))  # 이전 화면으로 되돌아가기.


from django.contrib.auth import get_user_model
from django.contrib import messages
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required


@login_required()
def student_password_reset(request, profile_id):
    '''학생의 계정 비밀번호 초기화.'''
    if request.method == 'POST':
        # 학생 프로필 가져오기
        student = get_object_or_404(models.Profile, id=profile_id)
        school = student.school
        # 교사 권한 확인
        if not check.Teacher(school=school, user=request.user).in_school():
            messages.error(request, "부정접근.")
            return JsonResponse({'success': True, 'message': '부정접근.'})
        try:
            user = student.admin
            new_password = str(random.randint(100000, 999999))  # 새로운 비밀번호 생성
            user.set_password(new_password)  # 비밀번호 설정
            user.save()  # 사용자 정보 저장
            # 성공 메시지
            return JsonResponse({'success': True, 'message': f'비밀번호가 성공적으로 초기화되었습니다. \n새 비밀번호: {new_password}'})
        except models.Profile.DoesNotExist:
            return JsonResponse({'success': True, 'message': "사용자를 찾을 수 없습니다."})

    return redirect(request.META.get('HTTP_REFERER', None))  # 이전 화면으로 되돌아가기.


def validate_teacher_password(request):
    '''모달 등 요청에서 교사의 패스워드 검증.'''
    if request.method == 'POST':
        password = request.POST.get('password')
        # 현재 로그인한 교사의 사용자 객체를 가져옵니다.
        user = request.user
        # 패스워드 검증
        if user.check_password(password):
            return JsonResponse({'valid': True})
        else:
            return JsonResponse({'valid': False})
    return JsonResponse({'valid': False}, status=400)

@login_required()
def delete_profile(request, type, profile_id, baseRoom_id):
    """학생 프로필 삭제 뷰"""
    if request.method != "POST":
        return HttpResponseForbidden("잘못된 접근 방식입니다.")
    if type == 'school':
        profile = get_object_or_404(models.Profile, id=profile_id)
        school = get_object_or_404(models.School, pk=baseRoom_id)
        if request.user != school.master:  # 그래도 다시 검증.
            messages.error(request, '이 기능은 관리자만이 가능합니다.')
            return check.Teacher(school=school).redirect_to_school()
        # 해당 학생, 교사 프로필 가져오기
        # 학생 삭제 수행
        profile.delete()
        return JsonResponse({'success': True, 'message': "학생이 삭제되었습니다."})
    elif type == 'homeroom':
        profile = get_object_or_404(models.Profile, id=profile_id)
        homeroom = get_object_or_404(models.Homeroom, pk=baseRoom_id)
        if check.Teacher(user=request.user, homeroom=homeroom).in_homeroom_and_none:
            pass
        else:
            return check.Teacher.redirect_to_homeroom()
        profile.homeroom.remove(homeroom)
        return JsonResponse({'success': True, 'message': "학생이 해당 학급에서 제거되었습니다.(학교엔 남아 있습니다. 완전 삭제는 관리자에게.)"})
    return redirect('school_report:student_assignment', type=type, baseRoom_id=baseRoom_id)
        #redirect(reverse('school_report:student_assignment', kwargs={'type': type, 'baseRoom_id': baseRoom_id}))


@login_required()
def student_code_input(request, school_id):
    '''학생 인증하기.'''
    school = get_object_or_404(models.School, pk=school_id)
    context = {'school': school}
    student = check.Student(user=request.user, school=school).in_school_and_none()
    if student == None:
        pass
    else:
        messages.error(request, '이미 이 기관에 인증되었습니다.' + str(student.student_code))
        return redirect('school_report:school_main', school_id=school_id)
    if request.method == 'POST':  # 포스트로 요청이 들어온다면...
        confirm_code = request.POST.get('code')
        try:
            student = models.Profile.objects.filter(school=school, confirm_code=confirm_code)[0]
            if student.obtained == True:
                messages.error(request, '이미 누군가 등록한 프로필입니다.')
                return render(request, 'school_report/school/student_code_input.html', context)
            # 이상 없으면 확인 페이지로 이동시킨다.
            return redirect('school_report:student_code_confirm', student_id=student.id)
        except Exception as e:
            messages.error(request, e)
            messages.error(request, '코드 해당계정이 없습니다.')
    return render(request, 'school_report/school/student_code_input.html', context)

@login_required()
def student_code_confirm(request, student_id):
    student = get_object_or_404(models.Profile, pk=student_id)
    context={'student':student}
    if request.method == 'POST':  # 포스트로 요청이 들어온다면...
        confirm_code = request.POST.get('code')
        if confirm_code == student.confirm_code:
            student.admin = request.user
            student.obtained = True
            student.confirm_code = None
            student.save()
            #request.user.student = student  # 계정에 등록. 이거 없어도 될듯. 필요없어진 기능.
            #request.user.save()
            messages.info(request, '인증에 성공하였습니다.')
            return redirect('school_report:school_main', school_id=student.school.id)
        else:
            messages.error(request, '코드가 안맞는데요;')
            return render(request, 'school_report/school/student_code_confirm.html', context)
    return render(request, 'school_report/school/student_code_confirm.html', context)

def meal_info(request, school_id):
    school = get_object_or_404(models.School, pk=school_id)
    context = {'school': school, 'today_int': int(datetime.datetime.today().strftime('%Y%m'))}
    # 급식정보
    if school.school_code:  # 학교코드가 있어야 진행.
        school_meal = SchoolMealsApi(ATPT_OFCDC_SC_CODE=school.education_office, SD_SCHUL_CODE=str(school.school_code))
        school_meal_data = school_meal.get_data()
        try:  # 개학 전엔 식사정보가 없어 None을 반환한다. 그럼 에러뜸.=
            sorted_data = sorted(school_meal_data, key=lambda x: datetime.datetime.strptime(x['MLSV_YMD'], "%Y%m%d"))
            # 표에 넣기 위해 항목별로 차례대로 넣는다.
            meal_data = {'일자': [], '식사': [], '메뉴': [], '칼로리': [], '영양정보': [], '재료정보': []}
            for i in sorted_data:
                date = i['MLSV_YMD']
                date = date[4:6] + '월' + date[6:] + '일'
                meal_data['일자'].append(date)  # 급식일자.
                meal_data['식사'].append(i['MMEAL_SC_NM'])  # 조,중,석식 분류.
                meal_data['메뉴'].append(i['DDISH_NM'])  # 메뉴
                meal_data['칼로리'].append(i['CAL_INFO'])  # 칼로리.
                meal_data['영양정보'].append(i['NTR_INFO'])  # 영양정보
                meal_data['재료정보'].append(i['ORPLC_INFO'])  # 재료정보
            context['meal_data'] = meal_data
        except Exception as e:
            print(f"[급식 데이터 처리 오류] {e}")
    return render(request, 'school_report/school/meal_info.html', context)

def lost_item_board(request, school_id):
    '''분실물게시판으로 가는 링크.'''
    school = get_object_or_404(models.School, pk=school_id)
    context = {}
    if school.is_lost_item_board_active:
        board = models.LostItemBoard.objects.get(school=school)
        teacher = check.Teacher(user=request.user, school=school).in_school_and_none()
        context['teacher_message'] = board.teacher_message
        context['teacher'] = teacher
        context['board'] = board
        context['school'] = school
        context['items_teacher_report'] = models.LostItem.objects.filter(board=board, is_report=True, status='lost').order_by('-created_at')  # 제보.
        context['items'] = models.LostItem.objects.filter(board=board, is_report=False, status='lost').order_by('-created_at')
    return render(request, 'school_report/school/boards/lost_item_board.html', context)

def lost_item_board_teachers_message(request, school_id):
    '''분실물 게시판에 남기는 교사의 말. 을 ajax로 받아 저장.'''
    school = models.School.objects.get(id=school_id)
    teacher = check.Teacher(user=request.user, school=school).in_school_and_none()
    if request.method == 'POST' and teacher:
        # Ajax로 전송된 'content' 값 받기
        content = request.POST.get('content')
        if content:
            # 받은 콘텐츠를 DB에 저장하거나 처리하는 코드 (예: 모델 저장)
            board = models.LostItemBoard.objects.get(school=school)
            board.teacher_message = content
            board.teacher = teacher
            board.save()
            # 성공 응답
            return JsonResponse({'status': 'success', 'message': '저장되었습니다!'})

        return JsonResponse({'status': 'error', 'message': '내용이 비어있습니다.'})

    return JsonResponse({'status': 'error', 'message': '잘못된 요청입니다.'})

@login_required
def lost_item_board_report_item(request, board_id, is_report=False):
    '''들어온 분실물 안내. is_report 여부로 학생이 올리는 것인지, 들어온 분실물 등록인지 구분한다.'''
    board = models.LostItemBoard.objects.get(pk=board_id)
    context = {}
    if request.method == 'POST':  # 포스트로 요청이 들어온다면... 글을 올리는 기능.
        form = SchoolLostItemForm(request.POST, request.FILES)  # 폼을 불러와 내용입력을 받는다.
        school = board.school
        profile = check.Teacher(user=request.user, school=school).in_school()
        if not profile:
            return redirect('school_report:lost_item_board', school_id=school.id)
        if form.is_valid():  # 문제가 없으면 다음으로 진행.
            item = form.save(commit=False)  # commit=False는 저장을 잠시 미루기 위함.(입력받는 값이 아닌, view에서 다른 값을 지정하기 위해)
            item.board = board
            item.author = profile
            if is_report:
                item.is_report = True
            item.save()
            return redirect('school_report:lost_item_board', school_id=school.id)
    else:  # 포스트 요청이 아니라면.. form으로 넘겨 내용을 작성하게 한다.
        form = SchoolLostItemForm()
    context['form'] = form  # 폼에서 오류가 있으면 오류의 내용을 담아 create.html로 넘긴다.
    return render(request, 'school_report/school/boards/lost_item_report_item.html', context)
@login_required
def lost_item_modify(request, item_id):
    '''분실물 글 수정.'''
    item = get_object_or_404(models.LostItem, pk=item_id)
    if request.user != item.author.admin:
        messages.error(request, '수정권한이 없습니다')
        return redirect('school_report:lost_item_board', school_id=item.board.school.id)
    if request.method == "POST":
        original_photo = item.photo_item  # 기존 사진 백업
        form = SchoolLostItemForm(request.POST, request.FILES, instance=item)  # 받은 내용을 객체에 담는다. instance에 제대로 된 걸 넣지 않으면 새로운 인스턴스를 만든다.
        if request.POST.get('delete_photo_item'):  # 이상하게 valid 안에 넣으면 안되더라..
            if item.photo_item:
                item.photo_item.delete(save=False)  # 기존 사진 실제 파일 시스템에서 삭제
                item.photo_item = None
        if form.is_valid():
            item = form.save(commit=False)  # 잠시 보류.
            # 사진을 새로 업로드하지 않았고, 삭제 요청도 없으면 기존 사진 유지
            if not request.FILES.get('photo_item') and not request.POST.get('delete_photo_item'):
                item.photo_item = original_photo
            item.save(update_fields=['where', 'when', 'description', 'photo_item', 'status'])  # 최종 저장.
            #item.save(update_fields=['where', 'when', 'description', 'photo_item', 'status'])  # 업데이트하면 다른 파일을 새로이 저장해버리는 문제 발생;;
            return redirect('school_report:lost_item_board', school_id=item.board.school.id)
    else:  # GET으로 요청된 경우.
        form = SchoolLostItemForm(instance=item)  # 해당 모델의 내용을 가져온다!
        # 태그를 문자열화 하여 form과 함께 담는다.
    context = {'form': form}
    return render(request, 'school_report/school/boards/lost_item_report_item.html', context)


@login_required
def lost_item_board_upload_claim_photo(request, item_id):
    '''분실물 찾아간 사람의 사진 올리기.'''
    lost_item = get_object_or_404(models.LostItem, pk=item_id)
    board = lost_item.board
    if request.user != lost_item.author.admin:
        return HttpResponseForbidden()
    if request.method == 'POST' and request.FILES.get('photo_claimed'):
        lost_item.photo_claimed.delete(save=False)  # 기존 사진 삭제.
        lost_item.photo_claimed = request.FILES['photo_claimed']
        lost_item.status = 'found'
        lost_item.save(update_fields=['photo_claimed', 'status'])
        return redirect('school_report:lost_item_board', school_id=board.school.id)
    return redirect('school_report:lost_item_board', school_id=board.school.id)
@login_required
def lost_item_board_found_item(request, board_id):
    '''찾은 물건들 두는 곳.'''
    board = get_object_or_404(models.LostItemBoard, pk=board_id)
    school = board.school
    context = {}
    context['board'] = board
    context['school'] = school
    context['items'] = models.LostItem.objects.filter(board=board, status='found').order_by('-created_at')
    return render(request, 'school_report/school/boards/lost_item_found.html', context)

def suggestion_board(request, school_id):
    '''건의 게시판으로 가는 링크.'''
    school = get_object_or_404(models.School, pk=school_id)
    context = {}
    if school.is_suggestion_board_active:
        board = models.SuggestionBoard.objects.get(school=school)
        teacher = check.Teacher(user=request.user, school=school).in_school_and_none()
        context['teacher_message'] = board.teacher_message
        context['teacher'] = teacher
        context['board'] = board
        context['school'] = school
        #context['items_teacher_report'] = models.LostItem.objects.filter(board=board, is_report=True)  # 제보.
        #context['items'] = models.LostItem.objects.filter(board=board, is_report=False)
    return render(request, 'school_report/school/boards/suggestion_board.html', context)
def suggestion_board_teachers_message(request, school_id):
    '''게시판에 남기는 교사의 말. 을 ajax로 받아 저장.'''
    school = models.School.objects.get(id=school_id)
    teacher = check.Teacher(user=request.user, school=school).in_school_and_none()
    if request.method == 'POST' and teacher:
        # Ajax로 전송된 'content' 값 받기
        content = request.POST.get('content')
        if content:
            # 받은 콘텐츠를 DB에 저장하거나 처리하는 코드 (예: 모델 저장)
            board = models.SuggestionBoard.objects.get(school=school)
            board.teacher_message = content
            board.teacher = teacher
            board.save()
            # 성공 응답
            return JsonResponse({'status': 'success', 'message': '저장되었습니다!'})

        return JsonResponse({'status': 'error', 'message': '내용이 비어있습니다.'})

    return JsonResponse({'status': 'error', 'message': '잘못된 요청입니다.'})
def profile_reset(request, profile_id):
    '''기존 계정을 잊어버린 경우, 연동을 끊게 해주기.'''
    profile = models.Profile.objects.get(pk=profile_id)
    profile.obtained = False
    profile.confirm_code = random.randint(100000, 999999)
    profile.save()
    return redirect(request.META.get('HTTP_REFERER', None))  # 이전 화면으로 되돌아가기.
