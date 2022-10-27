import datetime

from django.shortcuts import render, get_object_or_404, redirect, resolve_url, HttpResponse
from .. import models  # 모델 호출.
from ..forms import HomeroomForm, AnnouncementForm
from django.contrib import messages
from django.contrib.auth.decorators import login_required
import openpyxl
import random
@login_required()
def create(request, school_id):
    school = get_object_or_404(models.School, pk=school_id)
    context = {'school': school}
    if request.method == 'POST':  # 포스트로 요청이 들어온다면... 글을 올리는 기능.
        if school.code != request.POST.get('code'):  # 코드가 같아야 진행.
            messages.error(request, "학교 코드가 맞지 않습니다.")
            context['form'] = HomeroomForm(request.POST)
            return render(request, 'school_report/homeroom/create.html', context)
        form = HomeroomForm(request.POST)  # 폼을 불러와 내용입력을 받는다.
        if form.is_valid():  # 문제가 없으면 다음으로 진행.
            homeroom = form.save(commit=False)  # commit=False는 저장을 잠시 미루기 위함.(입력받는 값이 아닌, view에서 다른 값을 지정하기 위해)
            homeroom.master = request.user  # 추가한 속성 author 적용
            homeroom.school = school
            homeroom.save()
            return render(request, 'school_report/homeroom/main.html', context)  # 작성이 끝나면 작성한 글로 보낸다.
    else:  # 포스트 요청이 아니라면.. form으로 넘겨 내용을 작성하게 한다.
        form = HomeroomForm()
    context['form'] = form  # 폼에서 오류가 있으면 오류의 내용을 담아 create.html로 넘긴다.
    return render(request, 'school_report/homeroom/create.html', context)

def main(request, homeroom_id):
    homeroom = get_object_or_404(models.Homeroom, pk=homeroom_id)
    context ={'homeroom': homeroom}

    classroom_list = homeroom.classroom_set.all()
    context['classroom_list'] = classroom_list
    announcement_list = homeroom.announcement_set.all()
    context['announcement_list'] = announcement_list

    return render(request, 'school_report/homeroom/main.html', context)

from . import check
@login_required()
def assignment(request, homeroom_id):
    homeroom = get_object_or_404(models.Homeroom, pk=homeroom_id)
    context = {'homeroom': homeroom}
    if check.Check_teacher(request, homeroom.school).in_school_and_none():
        resistered = models.Student.objects.filter(homeroom=homeroom, obtained=True)
        context['resistered'] = resistered
        unresistered = models.Student.objects.filter(homeroom=homeroom, obtained=False)  # 등록 안한 사람만 반환.
        context['unresistered'] = unresistered
        return render(request, 'school_report/homeroom/assignment.html', context)
    else:
        messages.error(request, '당신은 학급의 관리자가 아닙니다.')
        return redirect('school_report:homeroom_main', homeroom_id=homeroom_id)

@login_required()
def download_excel_form(request, homeroom_id):
    homeroom = get_object_or_404(models.Homeroom, pk=homeroom_id)
    students = homeroom.student_set.all()  # 학급에 속한 학생.
    wb = openpyxl.Workbook()
    ws = wb.create_sheet('명단 form', 0)
    ws['A1'] = '수험코드(학번)'
    ws['B1'] = '이름'
    a = 'A'  # 번호쓰는 라인.
    b = 'B'  # 이름쓰는 라인.
    for i, student in enumerate(students):
        num = str(i + 2)
        ws[a + num] = student.student_code
        ws[b + num] = student.name
    response = HttpResponse(content_type="application/vnd.ms-excel")
    response["Content-Disposition"] = 'attachment; filename=' + '명단양식' + '.xls'
    wb.save(response)
    return response

@login_required()
def upload_excel_form(request, homeroom_id):
    context = {}
    if request.method == "POST":
        homeroom = get_object_or_404(models.Homeroom, pk=homeroom_id)
        context['homeroom'] = homeroom
        uploadedFile = request.FILES["uploadedFile"]  # post요청 안의 name속성으로 찾는다.
        wb = openpyxl.load_workbook(uploadedFile, data_only=True)  # 파일을 핸들러로 읽는다.
        work_sheet = wb["명단 form"]  # 첫번째 워크시트를 사용한다.

        # 엑셀 데이터를 리스트 처리한다.
        work_sheet_data = []  # 전체 데이터를 담기 위한 리스트.
        for row in work_sheet.rows:  # 열을 순회한다.
            row_data = []  # 열 데이터를 담기 위한 리스트
            for cell in row:
                row_data.append(cell.value)  # 셀 값을 하나씩 리스트에 담는다.
            work_sheet_data.append(row_data)  # 워크시트 리스트 안에 열 리스트를 담아...
            # work_sheet_data[열번호][행번호] 형태로 엑셀의 데이터에 접근할 수 있게 된다.
        work_sheet_data = work_sheet_data[1:]  # 첫번째 행은 버린다.
        teacher = check.Check_teacher(request, homeroom).in_homeroom_and_none()
        if teacher != None:
            for data in work_sheet_data:  # 행별로 데이터를 가져온다.
                student_code = data[0]
                name = data[1]
                student, created = models.Student.objects.get_or_create(school=homeroom.school,
                                                                        student_code=student_code)
                student.homeroom.add(homeroom)
                student.name = name
                if created:
                    student.code = random.randint(100000, 999999)  # 코드 지정.
                student.save()
            messages.info(request,'반영 완료.')

    return redirect('school_report:homeroom_student_assignment', homeroom_id=homeroom_id)

@login_required()
def announcement_create(request, homeroom_id):
    homeroom = get_object_or_404(models.Homeroom, pk=homeroom_id)
    context = {'homeroom':homeroom}
    if request.method == 'POST':  # 포스트로 요청이 들어온다면... 글을 올리는 기능.
        form = AnnouncementForm(request.POST)  # 폼을 불러와 내용입력을 받는다.
        if form.is_valid():  # 문제가 없으면 다음으로 진행.
            announcement = form.save(commit=False)  # commit=False는 저장을 잠시 미루기 위함.(입력받는 값이 아닌, view에서 다른 값을 지정하기 위해)
            announcement.author = request.user  # 추가한 속성 author 적용
            announcement.homeroom = homeroom  # 게시판 지정.
            announcement.save()

            # 개별 확인을 위한 개별공지 생성.
            student_list = models.Student.objects.filter(homeroom=homeroom)
            for student in student_list:
                individual, created = models.AnnoIndividual.objects.get_or_create(to_student=student,
                          base_announcement=announcement)

            return redirect('school_report:homeroom_main', homeroom.id)  # 작성이 끝나면 작성한 글로 보낸다.
    else:  # 포스트 요청이 아니라면.. form으로 넘겨 내용을 작성하게 한다.
        form = AnnouncementForm()
    context['form'] = form  # 폼에서 오류가 있으면 오류의 내용을 담아 create.html로 넘긴다.
    return render(request, 'school_report/homeroom/announcement_create.html', context)

def announcement_detail(request, posting_id):
    context = {}
    posting = get_object_or_404(models.Announcement, pk=posting_id)
    context['posting'] = posting
    homeroom = posting.homeroom
    context['homeroom'] = homeroom

    student = check.Check_student(request, homeroom).in_homeroom_and_none()
    # 학생과 교사 가르기.
    if student:
        # 새로운 학생이 훗날 추가되었다면 접속했을 때 개별공지 하나가 늘게끔.
        individual_announcement, created = models.AnnoIndividual.objects.get_or_create(to_student=student, base_announcement=posting)
        context['individual_announcement'] = individual_announcement
    elif posting.author == request.user:
        annoIndividual_list = models.AnnoIndividual.objects.filter(base_announcement=posting)
        context['annoIndividual_list'] = annoIndividual_list  # 이건 나중에 없애도 될듯.
    else:
        return check.Check_student(request, homeroom).redirect_to_homeroom()
    #individual_announcement = get_object_or_404(models.AnnoIndividual, )
    return render(request, 'school_report/homeroom/announcement/detail.html', context)

def announcement_modify(request, posting_id):
    posting = get_object_or_404(models.Announcement, pk=posting_id)
    if request.user != posting.author:
        messages.error(request, '수정권한이 없습니다')
        return redirect('school_report:announcement_detail', posting_id=posting.id)
    if request.method == "POST":
        form = AnnouncementForm(request.POST, instance=posting)  # 받은 내용을 객체에 담는다. instance에 제대로 된 걸 넣지 않으면 새로운 인스턴스를 만든다.
        if form.is_valid():
            posting = form.save(commit=False)
            posting.author = request.user
            posting.save()
            # 개별 확인을 위한 개별과제 체크 해제.
            submit_list = models.AnnoIndividual.objects.filter(base_announcement=posting)
            for submit in submit_list:
                submit.check = False
            return redirect('school_report:announcement_detail', posting_id=posting.id)
    else:  # GET으로 요청된 경우.
        form = AnnouncementForm(instance=posting)  # 해당 모델의 내용을 가져온다!
        # 태그를 문자열화 하여 form과 함께 담는다.
    context = {'form': form}
    messages.error(request, '수정하면 기존 확인한 학생들의 체크는 "읽지않음"으로 갱신됩니다.')
    return render(request, 'school_report/homeroom/announcement_create.html', context)

def announcement_delete(request, posting_id):
    posting = get_object_or_404(models.Announcement, pk=posting_id)
    if request.user != posting.author:
        messages.error(request, '삭제권한이 없습니다. 꼼수쓰지 마라;')
        return redirect('school_report:announcement_detail', posting_id=posting.id)
    messages.success(request, '삭제 성공~!')
    homeroom = posting.homeroom
    posting.delete()
    return redirect('school_report:homeroom_main', homeroom_id=homeroom.id)

def individual_download_excel_from(request, announcement_id):
    wb = openpyxl.Workbook()
    ws = wb.create_sheet('명단 form', 0)
    ws['A1'] = '번호'
    ws['B1'] = '이름'
    ws['C1'] = '공지내용'
    announcement = get_object_or_404(models.Announcement, pk=announcement_id)
    homeroom = announcement.homeroom
    student_list = models.Student.objects.filter(homeroom=homeroom)
    student_list = student_list.order_by('student_code')
    for i, student in enumerate(student_list):
        row = str(i+2)
        number_col = 'A' + row
        name_col = 'B' + row
        content_col = 'C' + row
        ws[number_col] = student.student_code
        ws[name_col] = student.name
        indianno = models.AnnoIndividual.objects.get(to_student=student, base_announcement=announcement)
        ws[content_col] = indianno.content

    response = HttpResponse(content_type="application/vnd.ms-excel")
    response["Content-Disposition"] = 'attachment; filename=' + '명단양식' + '.xls'
    wb.save(response)

    return response

def individual_upload_excel_form(request, announcement_id):
    context = {}
    if request.method == "POST":
        announcement = get_object_or_404(models.Announcement, pk=announcement_id)
        context['announcement'] = announcement
        uploadedFile = request.FILES["uploadedFile"]  # post요청 안의 name속성으로 찾는다.
        wb = openpyxl.load_workbook(uploadedFile, data_only=True)  # 파일을 핸들러로 읽는다.
        work_sheet = wb["명단 form"]  # 첫번째 워크시트를 사용한다.

        # 엑셀 데이터를 리스트 처리한다.
        work_sheet_data = []  # 전체 데이터를 담기 위한 리스트.
        for row in work_sheet.rows:  # 열을 순회한다.
            row_data = []  # 열 데이터를 담기 위한 리스트
            for cell in row:
                row_data.append(cell.value)  # 셀 값을 하나씩 리스트에 담는다.
            work_sheet_data.append(row_data)  # 워크시트 리스트 안에 열 리스트를 담아...
            # work_sheet_data[열번호][행번호] 형태로 엑셀의 데이터에 접근할 수 있게 된다.
        work_sheet_data = work_sheet_data[1:]  # 첫번째 행은 버린다.

        if request.user == announcement.author:
            for data in work_sheet_data:  # 행별로 데이터를 가져온다.
                name = str(data[1])
                content = str(data[2])
                homeroom = announcement.homeroom
                student = get_object_or_404(models.Student, name=name, homeroom=homeroom)
                individual, created = models.AnnoIndividual.objects.get_or_create(to_student=student, base_announcement=announcement)
                individual.content = content
                individual.save()

    return redirect('school_report:announcement_detail', posting_id=announcement_id)  # 필요에 따라 렌더링.
    pass##엑셀파일 받아서...

def announcement_check(request, announcement_id):
    '''개별공지 확인체크.'''
    announcement = get_object_or_404(models.AnnoIndividual, pk=announcement_id)
    student_owner = announcement.to_student.admin
    if request.user == student_owner:  # 당사자만 체크 가능.
        announcement.check = True
        announcement.check_date = datetime.datetime.now()
        announcement.save()
        messages.info(request, '공지 확인하였습니다~')
    else:
        messages.error(request, '본인의 공지만 체크할 수 있습니다.')
    return redirect('school_report:announcement_detail', posting_id=announcement.base_announcement.id)