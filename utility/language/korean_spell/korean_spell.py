from django.shortcuts import render, get_object_or_404, redirect, resolve_url
import openpyxl
import requests
from bs4 import BeautifulSoup
import json

from utility.tasks import correct_text
from django.contrib.auth.decorators import login_required
from utility.models import Spell, SpellObject
from config import secret

def main(request):
    context = {}
    # 검사객체 올리기. 아래 함수 안에 있는 것과 동일.
    try:
        spell_user = Spell.objects.get(user=request.user)  # 위에서 객체가 만들어지기 전에 넘어갈 수 있어 try처리.
        spell_objects = SpellObject.objects.filter(spell=spell_user)  # 맞춤법 객체들을 모은다.
        context['spell_user'] = spell_user
        context['spell_objects'] = spell_objects
    except Exception as e:
        print(e)
    return render(request, 'utility/korean_spell/main.html', context)

def han_spell(text):
    # 0. 수정 데이터 담기.
    cor_num = 0  # 수정한 횟수.
    # 1. 텍스트 준비 & 개행문자 처리
    text = text.replace('\n', '\r\n')  # 개행문자를 검사기에 맞게 변경.
    ### 이젠 chkKey를 동적으로 생성해서 자동 접근을 막은 듯하다... 해결방법은 찾지 못함;
    with requests.Session() as session:
        # 2. 페이지 로드하여 chkKey 추출
        response = session.get('https://nara-speller.co.kr/speller/')
        soup = BeautifulSoup(response.text, 'html.parser')
        chk_key = soup.find('input', {'id': 'chkKey'}).get('value')

        # 3. 맞춤법 검사 요청
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3',
            'Content-Type': 'application/x-www-form-urlencoded',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8'
        }
        data = {
            'text1': text,
            'chkKey': chk_key
        }
        response = session.post('https://nara-speller.co.kr/speller/results', data=data, headers=headers)
        return response.text
    print(response)
    # 3. 응답에서 필요한 내용 추출 (html 파싱)
    data = response.text.split('data = [', 1)[-1].rsplit('];', 1)[0]  # 단순 데이터 찾기..
    try:
        # 4. 파이썬 딕셔너리 형식으로 변환
        data = json.loads(data)
    except:
        corrected_text = '없음.'
        return cor_num, corrected_text  # 에러가 생기면 바로 반환해버린다.(완벽하면 None이 반환됨.)
    # 5. 교정 내용 출력
    correction_datas = data['errInfo']
    corrected_text = text  # 흐흠, 어째서인지 이렇게 해도 복사가 된다. 편하네.
    pointer = 0  # 기본 개념과는 다르지만... 문자열을 삭제하고 나면 이동해야 할 포인터.
    # print(correction_datas)
    for correction_data in correction_datas:
        # 정보 받기.
        start = pointer + correction_data['start']  # 원문의 어디에서 시작하는지.
        end = pointer + correction_data['end']  # 원문의 어디에서 끝나는지.
        target = correction_data['orgStr']  # 문제가 있는 부분.
        help = correction_data['help']  # 무엇이 잘못된 것인지.
        changed = correction_data['candWord']  # 고쳐진 말.
        changed = '<span class="text-bg-danger opacity-75" title="{}">{}</span>'.format(help, changed)  # html처리.
        # 조작.
        # print(correction_data['orgStr'])
        corrected_text = corrected_text[:start] + changed + corrected_text[end:]  # 잘못된 용어 제거 + 교정어 삽입
        erased_len = len(changed) - len(target)
        pointer = pointer + erased_len  # 지워진 문자 만큼 포인터를 옮겨준다. 다음 문자열 제거를 위해서.
        # 문제정보 담기.
        cor_num += 1  # 수정횟수 증가.
    return cor_num, corrected_text


@login_required()
def upload_excel_form(request):
    context = {}
    if request.method == "POST":
        uploadedFile = request.FILES["uploadedFile"]  # post요청 안의 name속성으로 찾는다.
        wb = openpyxl.load_workbook(uploadedFile, data_only=True)  # 파일을 핸들러로 읽는다.
        sheetnames = wb.sheetnames
        work_sheet = wb[sheetnames[0]]  # 첫번째 워크시트를 사용한다.

        # 엑셀 데이터를 처리한다. 리스트에 담는다.
        inner_data = []
        for row in work_sheet.rows:  # 열을 순회한다.
            for cell in row:  # 줄 안의 셀을 순회한다.
                cell = cell.value
                if cell == None:
                    continue  # 데이터가 없으면 넘기기.
                inner_data.append(str(cell))
        list_data = json.dumps(inner_data)  # json으로 바꾼다.

        if secret.Is_server:
            correct_text.delay(request.user.id, list_data)  # 비동기처리.
        else:
            print(list_data)
            correct_text(request.user.id, list_data)
        return redirect('utility:korean_spell')


    else:  # 하나씩 처리할 때. GET으로 받는다.
        wrong_text = {}  # 잘못된 정보를 담을 사전.
        check_text = request.GET["check_text"]
        cor_num, corrected_text = han_spell(check_text)
        wrong_text[check_text] = corrected_text
        context['correct_info'] = wrong_text
        # 검사객체 올리기. main 함수 안에 있는 것과 동일.
        try:
            spell_user = Spell.objects.get(user=request.user)  # 위에서 객체가 만들어지기 전에 넘어갈 수 있어 try처리.
            spell_objects = SpellObject.objects.filter(spell=spell_user)  # 맞춤법 객체들을 모은다.
            context['spell_user'] = spell_user
            context['spell_objects'] = spell_objects
        except Exception as e:
            print(e)
    return render(request, 'utility/korean_spell/main.html', context)